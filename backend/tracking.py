"""
tracking.py — Sincronización del historial de revisiones con un Excel en OneDrive.

Flujo:
- Al arrancar la app: pull_from_onedrive() → restaura estado en SQLite local
- Al cambiar estado (aprobar/rechazar/skip): push_to_onedrive() → actualiza Excel
"""

import io
import os
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import requests
from dotenv import load_dotenv

from sharepoint import get_session, download_file, upload_file, ONEDRIVE_BASE_URL, ONEDRIVE_ROOT_FOLDER
from database import get_conn

load_dotenv(Path(__file__).parent.parent / ".env")

TRACKING_FILENAME = "TRACKING_BITACORAS_2026.xlsx"

def _onedrive_full_path() -> str:
    """Ruta server-relative completa: /personal/usuario/Documents/Instructor Luis Eduardo"""
    user = ONEDRIVE_BASE_URL.split("/personal/")[1]
    return f"/personal/{user}{ONEDRIVE_ROOT_FOLDER}"

TRACKING_SERVER_PATH = f"{_onedrive_full_path()}/{TRACKING_FILENAME}"

COLUMNS = [
    "id", "ficha", "cc", "apprentice_name", "program", "area", "company",
    "start_date", "end_date", "bitacora_num", "filename", "file_path",
    "file_ext", "modified", "status", "notes", "reviewed_at", "signed_pdf_path",
    "first_seen", "last_synced"
]

STATUS_COLORS = {
    "pending":  "FFF3E0",
    "approved": "E8F5E9",
    "rejected": "FFEBEE",
    "skipped":  "E3F2FD",
}


def _build_workbook(rows: list) -> openpyxl.Workbook:
    """Construye el Excel de tracking desde la lista de registros."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracking Bitácoras"

    # Header — usar los key names para que el pull los encuentre directamente
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, key in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row_idx, row in enumerate(rows, 2):
        status = row["status"] if hasattr(row, "__getitem__") else getattr(row, "status", "pending")
        fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
        for col_idx, key in enumerate(COLUMNS, 1):
            val = row[key] if hasattr(row, "__getitem__") else getattr(row, key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill

    # Anchos de columna
    widths = {1:6, 2:12, 3:14, 4:35, 5:40, 6:16, 7:35, 8:12, 9:12,
              10:10, 11:45, 12:60, 13:6, 14:22, 15:12, 16:40, 17:20, 18:60, 19:20, 20:20}
    for col, width in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Freeze header
    ws.freeze_panes = "A2"
    return wb


def push_to_onedrive(session: requests.Session = None):
    """
    Lee todos los registros del SQLite local y sube el Excel actualizado a OneDrive.
    """
    if session is None:
        session = get_session()

    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM bitacoras ORDER BY ficha, apprentice_name, bitacora_num").fetchall()

    wb = _build_workbook(rows)

    out = io.BytesIO()
    wb.save(out)
    content = out.getvalue()

    ok = upload_file(session, ONEDRIVE_BASE_URL, _onedrive_full_path(), TRACKING_FILENAME, content)
    if ok:
        print(f"✓ Tracking subido a OneDrive: {TRACKING_SERVER_PATH}")
    else:
        print(f"✗ Error subiendo tracking a OneDrive")
    return ok


def pull_from_onedrive(session: requests.Session = None):
    """
    Descarga el Excel de tracking desde OneDrive y restaura el estado en SQLite.
    Solo actualiza status, notes, reviewed_at, signed_pdf_path — no toca metadatos.
    Retorna número de registros restaurados.
    """
    if session is None:
        session = get_session()

    content = download_file(session, ONEDRIVE_BASE_URL, TRACKING_SERVER_PATH)
    if not content:
        print("No existe tracking en OneDrive aún — se creará al primer cambio de estado.")
        return 0

    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Leer headers de la primera fila
    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {name: idx for idx, name in enumerate(header_row) if name}

    restored = 0
    with get_conn() as conn:
        for row_idx in range(2, ws.max_row + 1):
            def v(name):
                idx = col.get(name)
                return ws.cell(row_idx, idx).value if idx else None

            file_path = v("file_path")
            status = v("status")
            if not file_path or not status:
                continue

            # Solo actualizar campos de revisión, no metadatos
            conn.execute("""
                UPDATE bitacoras
                SET status = ?, notes = ?, reviewed_at = ?, signed_pdf_path = ?
                WHERE file_path = ?
            """, (status, v("notes"), v("reviewed_at"), v("signed_pdf_path"), file_path))
            restored += 1

    print(f"✓ Tracking restaurado desde OneDrive: {restored} registros")
    return restored
